import json
import os
from PyQt5.QtWidgets import (QFileDialog, QMessageBox ,QTreeWidgetItem)
from PyQt5.QtGui import QPixmap, QIcon
from PIL import Image, ImageFile
from openpyxl import load_workbook
from fpdf import FPDF
import shutil
import warnings
from datetime import datetime
warnings.filterwarnings("ignore", message="cmap value too big/small")

Image.MAX_IMAGE_PIXELS = None  # Disable the limit on image size
ImageFile.LOAD_TRUNCATED_IMAGES = True  # Enable loading of truncated images

class PDF(FPDF):
    def header(self):
        self.set_font('SimHei', '', 12)
        self.cell(0, 10, '郑州大学第三附属医院 河南省妇幼保健院', 0, 1, 'C')
        self.set_font('SimHei', '', 12)
        self.cell(0, 10, '病理检查报告单', 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('SimHei', 'I', 8)
        self.cell(0, 10, f'第 {self.page_no()} 页', 0, 0, 'C')

    def patient_info(self, data):
        self.set_font('SimHei', '', 12)
        self.cell(0, 10, f"病理号: {data['patient_index']}", 0, 1)
        self.cell(0, 10, f"姓名: {data['name']}", 0, 1)
        self.cell(0, 10, f"性别: {data['gender']}", 0, 1)
        self.cell(0, 10, f"年龄: {data['age']} 岁", 0, 1)
        self.cell(0, 10, f"送检科室: {data['department']}", 0, 1)
        self.cell(0, 10, f"送检日期: {data['submission_time']}", 0, 1)
        self.cell(0, 10, f"住院号: {data['admission_number']}", 0, 1)
        self.cell(0, 10, f"床号: {data['bed_number']}", 0, 1)
        self.cell(0, 10, f"送检医生: {data['doctor']}", 0, 1)
        self.cell(0, 10, f"临床诊断: {data['result']}", 0, 1)
        self.ln(10)

    def add_images(self, folder):
        for image_file in os.listdir(folder):
            if image_file.endswith(".jpg") or image_file.endswith(".png"):
                self.image(os.path.join(folder, image_file), x=10, w=180)
                self.ln(10)

    def add_diagnosis(self, text):
        self.set_font('SimHei', '', 12)
        self.multi_cell(0, 10, text)


def update_patient_info(self):
    self.current_patient_data['patient_index'] = self.patient_index_edit.text() # 病理号
    self.current_patient_data['name'] = self.name_edit.text() # 姓名
    self.current_patient_data['gender'] = self.gender_edit.text() # 性别
    self.current_patient_data['age'] = int(self.age_edit.text()) # 年龄
    self.current_patient_data['submission_time'] = self.submission_time_edit.text() # 送检时间
    self.current_patient_data['department'] = self.department_edit.text() # 送检科室
    self.current_patient_data['admission_number'] = self.admission_number_edit.text() # 住院号
    self.current_patient_data['bed_number'] = self.bed_number_edit.text() # 床号
    self.current_patient_data['doctor'] = self.doctor_edit.text() # 送检医生
    self.current_patient_data['result'] = self.result_edit.text() # 临床诊断
    
    patient_folder = os.path.join(self.data_folder, self.current_patient_data['folder'])
    info_path = os.path.join(patient_folder, 'info.json')
    
    with open(info_path, 'w', encoding='utf-8') as f:
        json.dump(self.current_patient_data, f, ensure_ascii=False, indent=4)
    
    # 获取当前选中的项
    current_item = self.patient_list.currentItem()

    # 判断当前选中的项是否为根目录或者子目录
    if current_item:
        # 如果当前选中的项有父项，则其父项为根目录
        if current_item.parent():
            root_item = current_item.parent()
        else:
            # 如果当前选中的项没有父项，则其本身为根目录
            root_item = current_item
        
        # 更新根目录的文本
        if root_item:
            root_item.setText(0, self.current_patient_data['name'] + '(' + self.current_patient_data['patient_index'] + ')')

    QMessageBox.information(self, '信息', '患者基本信息已更新。')

def update_diagnosis_report(self):
    self.current_patient_data['diagnosis_report'] = self.report_text.toPlainText()
    #self.current_patient_data['diagnosis_report'] = self.pdf_diagnosis_text
    
    patient_folder = os.path.join(self.data_folder, self.current_patient_data['folder'])
    info_path = os.path.join(patient_folder, 'info.json')
    
    with open(info_path, 'w', encoding='utf-8') as f:
        json.dump(self.current_patient_data, f, ensure_ascii=False, indent=4)
    
    # PDF生成逻辑
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font('SimHei', '', 'res/simhei.ttf', uni=True)

    # Title
    pdf.set_font("SimHei", size=16)
    pdf.cell(200, 10, txt="郑州大学第三附属医院 河南省妇幼保健院", ln=True, align='C')
    pdf.set_font("SimHei", size=14)
    pdf.cell(200, 10, txt="病理诊断报告单", ln=True, align='C')
    pdf.ln(1)

    # Patient Info
    pdf.set_font("SimHei", size=12)
    pdf.cell(0, 10, txt=f"病理号: {self.current_patient_data['patient_index']}", ln=True, align='R')
    pdf.set_line_width(0.5)
    pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
    pdf.ln(1)

    pdf.cell(0, 10, txt=f"姓名: {self.current_patient_data['name']}            性别: {self.current_patient_data['gender']}            年龄: {self.current_patient_data['age']} 岁            送检科室: {self.current_patient_data['department']}", ln=True)
    pdf.cell(0, 10, txt=f"住院号: {self.current_patient_data['admission_number']}                 床号: {self.current_patient_data['bed_number']}                送检医生: {self.current_patient_data['doctor']}", ln=True)
    pdf.cell(0, 10, txt=f"临床诊断: {self.current_patient_data['result']}                    送检日期: {self.current_patient_data['submission_time']}", ln=True)
    pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
    pdf.ln(1)

    # 添加光镜所见
    pdf.cell(0, 10, txt="光镜所见:", ln=True)
    pdf.ln(1)

    # 横向排列两个图片
    images = [img for img in os.listdir(patient_folder) if img.endswith(".jpg") or img.endswith(".png")]

    if len(images) >= 2:
        image1_path = os.path.join(patient_folder, images[0])
        image2_path = os.path.join(patient_folder, images[1])
        image_width = 80
        margin = (210 - (2 * image_width)) / 3  # A4宽度为210，计算左右边距和图片间距
        x1 = margin
        x2 = margin * 2 + image_width
        pdf.image(image1_path, x=x1, y=pdf.get_y(), w=image_width)  # 缩放图片1
        pdf.image(image2_path, x=x2, y=pdf.get_y(), w=image_width)  # 缩放图片2
        pdf.ln(90)  # 调整行间距，根据图片的高度调整
    
    # Diagnosis report
    pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
    pdf.ln(1)
    pdf.set_font("SimHei", size=12)
    pdf.multi_cell(0, 10, txt=f"{self.pdf_text}")
    pdf.ln(50)

    pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
    pdf.ln(1)
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pdf.cell(0, 10, txt=f"诊断医师:                  复诊医师:                 报告日期: {current_datetime}", ln=True)

    # Save the PDF
    pdf_output_path = os.path.join(patient_folder, 'diagnosis_report.pdf')
    pdf.output(pdf_output_path)
    
    QMessageBox.information(self, '信息', '诊断报告已更新。')

def generate_patient_id(self):
    existing_ids = [int(folder) for folder in os.listdir(self.data_folder) if folder.isdigit()]
    if existing_ids:
        new_id = max(existing_ids) + 1
    else:
        new_id = 1
    return f"{new_id:04d}"


def add_new_patient(self):
    new_patient_name = generate_patient_id(self)
    new_patient_folder = os.path.join(self.data_folder, new_patient_name)
    if not os.path.exists(new_patient_folder):
        os.makedirs(new_patient_folder)
    new_patient_info = {
        'patient_index': 000000,
        'name': new_patient_name,
        'gender': '',
        'age': 0,
        'department': '',
        'submission_time': '',
        'admission_number': '',
        'bed_number': '',
        'doctor': '',
        'result': '',
        'diagnosis_report': '',
        'folder': new_patient_name
    }
    info_path = os.path.join(new_patient_folder, 'info.json')
    with open(info_path, 'w', encoding='utf-8') as f:
        json.dump(new_patient_info, f, ensure_ascii=False, indent=4)
    self.patients.append(new_patient_info)

    patient_item = QTreeWidgetItem([f"{new_patient_info['name']}({new_patient_info['patient_index']})"])
    
    patient_item.setIcon(0, QIcon("res/patient_2.png"))  # 设置母目录图标
    self.patient_list.addTopLevelItem(patient_item)
    self.clear_patient_info()
    QMessageBox.information(self, '信息', f'新患者 {new_patient_name} 已添加。')


def import_new_image_group(self):
    options = QFileDialog.Options()
    folder_name = QFileDialog.getExistingDirectory(self, "选择文件夹", options=QFileDialog.ShowDirsOnly)
    if folder_name:
        for patient_id in os.listdir(folder_name):
            patient_subfolder = os.path.join(folder_name, patient_id)
            if os.path.isdir(patient_subfolder):
                for subdir_name in os.listdir(patient_subfolder):
                    subdir_path = os.path.join(patient_subfolder, subdir_name)
                    for j in range(len(self.patients)):
                        current_patient_id = self.patients[j]['patient_index']
                        if str(current_patient_id) == patient_id:
                            selected_items = j
                            if selected_items is not None:
                                patient_folder = os.path.join(self.data_folder, self.patients[selected_items]['folder'])
                                new_image_name = f'image{len([name for name in os.listdir(patient_folder) if os.path.isdir(os.path.join(patient_folder, name))]) + 1}'
                                new_image_path = os.path.join(patient_folder, new_image_name)
                                shutil.copytree(subdir_path, new_image_path)
                                resized_image_old_path = os.path.join(new_image_path, f'{subdir_name}_resized.png')
                                resized_image_new_path = os.path.join(new_image_path, f'{new_image_name}_resized.png')
                                if os.path.exists(resized_image_old_path):
                                    os.rename(resized_image_old_path, resized_image_new_path)
                                try:
                                    self.image_folders.append(new_image_name)
                                except:
                                    print('ok')

                                
                                # 更新子目录
                                for i in range(self.patient_list.topLevelItemCount()):
                                    item = self.patient_list.topLevelItem(i)
                                    if item.text(0).startswith(self.patients[selected_items]['name']):
                                        new_image_item = QTreeWidgetItem([new_image_name])
                                        new_image_item.setIcon(0, QIcon("res/slice.png"))  # 设置子目录图标
                                        item.addChild(new_image_item)
                                        item.setExpanded(True)
                                        # 设置母目录图标
                                        item.setIcon(0, QIcon("res/patient.png"))
                                        break
        QMessageBox.information(self, '信息', f'新切片已批量导入。')
        root = self.patient_list.invisibleRootItem()
        for i in range(root.childCount()):
                item = root.child(i)
                item.setExpanded(False)


from openpyxl import load_workbook
import os
import json
from PyQt5.QtWidgets import QFileDialog, QTreeWidgetItem
from PyQt5.QtGui import QIcon

def add_new_patient_group(self):
    # 打开文件对话框，选择 Excel 文件
    file_path, _ = QFileDialog.getOpenFileName(self, 'Open file', 'D:\\Yuhang\\', 'Excel files (*.xlsx)')
    
    if file_path:
        # 使用 openpyxl 加载 Excel 文件
        wb = load_workbook(filename=file_path)
        sheet = wb.active  # 获取活动工作表

        # 读取前四列数据
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取（假设第一行是表头）
            # 提取前四列的数据
            patient_index, name, gender, age = row[:4]

            # 生成新的患者 ID 和文件夹
            new_patient_name = generate_patient_id(self)
            new_patient_folder = os.path.join(self.data_folder, new_patient_name)
            if not os.path.exists(new_patient_folder):
                os.makedirs(new_patient_folder)

                # 创建患者信息字典
                new_patient_info = {
                    'patient_index': patient_index,
                    'name': name,
                    'gender': gender,
                    'age': age,
                    'department': '',
                    'submission_time': '',
                    'admission_number': '',
                    'bed_number': '',
                    'doctor': '',
                    'result': '',
                    'diagnosis_report': '',
                    'folder': new_patient_name
                }

                # 将患者信息保存为 JSON 文件
                info_path = os.path.join(new_patient_folder, 'info.json')
                with open(info_path, 'w', encoding='utf-8') as f:
                    json.dump(new_patient_info, f, ensure_ascii=False, indent=4)

                # 将患者信息添加到列表中
                self.patients.append(new_patient_info)

                # 更新 QTreeWidget
                patient_item = QTreeWidgetItem([f"{new_patient_info['name']}({new_patient_info['patient_index']})"])
                patient_item.setIcon(0, QIcon("res/patient_2.png"))  # 设置图标
                self.patient_list.addTopLevelItem(patient_item)  # 添加为顶层项

        # 清空患者信息显示
        self.clear_patient_info()



def clear_patient_info(self):
    self.patient_index_edit.clear()
    self.name_edit.clear()
    self.gender_edit.clear()
    self.age_edit.clear()
    self.department_edit.clear()
    self.submission_time_edit.clear()
    self.admission_number_edit.clear()
    self.bed_number_edit.clear()
    self.doctor_edit.clear()
    self.result_edit.clear()

    self.detailed_info_text.clear()
    self.report_text.clear()
    self.graphics_view_top.scene.clear()
    self.graphics_view_bottom.scene.clear()
    self.image_item_top = None
    self.image_item_bottom = None
    self.zoom_slider_top.setValue(10)
    self.zoom_slider_bottom.setValue(10)

def import_new_image(self):
    selected_items = self.patient_list.selectedItems()
    
    
    if selected_items:
        selected_item = selected_items[0]
        if selected_item.parent():
            selected_item = selected_item.parent()
        index = self.patient_list.indexOfTopLevelItem(selected_item)
        patient_folder = os.path.join(self.data_folder, self.patients[index]['folder'])
        options = QFileDialog.Options()
        folder_name = QFileDialog.getExistingDirectory(self, "选择文件夹", options=QFileDialog.ShowDirsOnly)
        if folder_name:
            new_image_name = f'image{len(self.image_folders) + 1}'
            new_image_path = os.path.join(patient_folder, new_image_name)
            last_dir_name = os.path.basename(folder_name)
            shutil.copytree(folder_name, new_image_path)
            resized_image_old_path = os.path.join(new_image_path, f'{last_dir_name}_resized.png')
            resized_image_new_path = os.path.join(new_image_path, f'{new_image_name}_resized.png')
            if os.path.exists(resized_image_old_path):
                os.rename(resized_image_old_path, resized_image_new_path)
            self.image_folders.append(new_image_name)
            
            # 更新子目录
            new_image_item = QTreeWidgetItem([new_image_name])
            new_image_item.setIcon(0, QIcon("res/slice.png"))  # 设置子目录图标
            selected_item.addChild(new_image_item)
            selected_item.setExpanded(True)

            # 设置母目录图标
            selected_item.setIcon(0, QIcon("res/patient.png"))

            self.display_image_top(os.path.join(new_image_path, f'{new_image_name}_resized.png'))
            QMessageBox.information(self, '信息', f'新切片已导入 {new_image_name}。')

def delete_patient(self):
    selected_items = self.patient_list.selectedItems()
    if selected_items:
        item = selected_items[0]
        if item.parent():
            item = item.parent()
        patient_name = item.text(0)
        print(f"Selected patient: {patient_name}")  # 调试输出
        print(f"All patients: {self.patients}")
        
        # 提取 patient_index 和 name 进行比较
        for index, patient_data in enumerate(self.patients):
            display_name = f"{patient_data['name']}({patient_data['patient_index']})"
            if display_name == patient_name:
                print(f"Found patient data: {patient_data}")  # 调试输出
                patient_folder = os.path.join(self.data_folder, patient_data['folder'])
                
                # Confirm deletion
                reply = QMessageBox.question(self, '确认删除', f'确定要删除患者 {patient_data["name"]} 的全部信息吗？', 
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    print("Confirmed deletion")  # 调试输出
                    # Remove the patient folder
                    if os.path.exists(patient_folder):
                        for root, dirs, files in os.walk(patient_folder, topdown=False):
                            for name in files:
                                os.remove(os.path.join(root, name))
                            for name in dirs:
                                os.rmdir(os.path.join(root, name))
                        os.rmdir(patient_folder)

                    # Remove from the QTreeWidget and the list
                    index_in_tree = self.patient_list.indexOfTopLevelItem(item)
                    if index_in_tree != -1:
                        self.patient_list.takeTopLevelItem(index_in_tree)
                    del self.patients[index]

                    QMessageBox.information(self, '信息', f'患者 {patient_data["name"]} 的全部信息已删除。')
                else:
                    print("Deletion cancelled")  # 调试输出
                break
        else:
            print("Patient data not found")  # 调试输出
    else:
        print("No item selected")  # 调试输出
